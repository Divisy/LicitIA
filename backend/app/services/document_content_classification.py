"""Classify SECOP documents by file content when filename rules fail (US 1.2.5 MVP: presupuesto)."""
from __future__ import annotations

import re
import shutil
import tempfile
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

from sqlalchemy.orm import Session

from app.config import settings
from app.core.logging import get_logger
from app.models.tender import Tender
from app.models.tender_document import TenderDocument
from app.services.document_storage import DocumentStorageService
from app.services.secop_document_filters import DocumentType
from app.services.secop_documents import (
    SecopDocumentDTO,
    build_document_object_key,
    build_document_storage_path,
    download_document_file,
    fetch_all_documents_for_portfolio,
    is_archive_filename,
)

logger = get_logger(__name__)

SUPPORTED_EXTENSIONS: frozenset[str] = frozenset({".pdf", ".xlsx", ".xls", ".xlsm"})

# Strong phrases that must appear in document body (not filename alone).
PRESUPUESTO_CONTENT_ANCHORS: tuple[str, ...] = (
    "presupuesto oficial",
    "formulario 1",
    "formulario no. 1",
    "formulario no 1",
    "formul1 presupuesto",
    "formul1",
    "analisis de precios unitarios",
    "análisis de precios unitarios",
    "analisis de precios",
    "análisis de precios",
)

FILENAME_EXCLUSION_KEYWORDS: tuple[str, ...] = (
    "matriz",
    "certificado",
    "autorizacion",
    "autorización",
    "experiencia",
    "indicadores",
    "vigencias futuras",
    "vigencia futura",
    "pda",
    "acreditacion",
    "acreditación",
    "conformacion",
    "conformación",
    "capacidad financiera",
    "convocatoria",
    "resolucion",
    "resolución",
    "acta",
    "aviso",
    "estudio previo",
    "viabilidad",
    "puntaje",
    "veedores",
    "cdp",
    "memoria justificativa",
    "carta de",
    "informe de evaluacion",
    "informe de evaluación",
)

CONTENT_EXCLUSION_PHRASES: tuple[str, ...] = (
    "certificado de disponibilidad presupuestal",
    "certificado de disponibilidad",
    "autorizacion de vigencias futuras",
    "autorización de vigencias futuras",
    "matriz de experiencia",
    "indicadores financieros",
    "capacidad financiera y organizacional",
    "aviso de convocatoria",
    "convocatoria publica",
    "convocatoria pública",
    "estudio previo",
    "viabilidad",
    "pliego de condiciones",
    "anexo tecnico",
    "anexo técnico",
    "acta de apertura",
    "memoria justificativa",
    "acreditacion de emprendimiento",
    "acreditación de emprendimiento",
)

MIN_PRESUPUESTO_AMOUNT = 100_000


def normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _parse_number(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace("$", "").replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    else:
        cleaned = cleaned.replace(",", ".")
    cleaned = re.sub(r"[^\d.-]", "", cleaned)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def is_excluded_presupuesto_filename(file_name: str) -> bool:
    haystack = normalize_text(file_name)
    if any(keyword in haystack for keyword in FILENAME_EXCLUSION_KEYWORDS):
        return True
    if re.search(r"\bformato\s+[2-9]\b", haystack):
        return True
    if re.search(r"\bformato\s+1[0-9]\b", haystack):
        return True
    return False


def is_excluded_presupuesto_content(text: str) -> bool:
    normalized = normalize_text(text)[: settings.PRESUPUESTO_CONTENT_CLASSIFICATION_MAX_CHARS]
    return any(phrase in normalized for phrase in CONTENT_EXCLUSION_PHRASES)


def content_has_presupuesto_anchor(text: str) -> bool:
    normalized = normalize_text(text)[: settings.PRESUPUESTO_CONTENT_CLASSIFICATION_MAX_CHARS]
    if not normalized:
        return False
    return any(phrase in normalized for phrase in PRESUPUESTO_CONTENT_ANCHORS)


def looks_like_presupuesto_workbook(path: Path) -> bool:
    """Require Formulario 1 / presupuesto oficial labels plus a large monetary total."""
    from openpyxl import load_workbook

    found_anchor = False
    found_amount = False

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        for sheet in workbook.worksheets[:5]:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= 120:
                    break
                cells = [cell for cell in row if cell is not None]
                if not cells:
                    continue
                joined = normalize_text(" ".join(str(cell) for cell in cells[:8]))
                if any(anchor in joined for anchor in PRESUPUESTO_CONTENT_ANCHORS):
                    found_anchor = True
                for cell in cells:
                    number = _parse_number(cell)
                    if number is not None and number >= MIN_PRESUPUESTO_AMOUNT:
                        found_amount = True
                        break
    finally:
        workbook.close()

    return found_anchor and found_amount


def classify_presupuesto_candidate(
    file_name: str,
    content_text: str,
    local_path: Optional[Path] = None,
    extension: str = "",
) -> bool:
    """Return True only for high-confidence presupuesto documents."""
    if is_excluded_presupuesto_filename(file_name):
        return False
    if is_excluded_presupuesto_content(content_text):
        return False
    if not content_has_presupuesto_anchor(content_text):
        return False

    ext = extension.lower()
    if ext in {".xlsx", ".xls", ".xlsm"}:
        if local_path is None or not local_path.is_file():
            return False
        return looks_like_presupuesto_workbook(local_path)

    if ext == ".pdf":
        return True

    return False


def candidate_extension(document: SecopDocumentDTO) -> str:
    extension = (document.extension or "").strip().lower()
    if extension and not extension.startswith("."):
        extension = f".{extension}"
    if extension:
        return extension
    return Path(document.file_name).suffix.lower()


def extract_text_from_local_pdf(path: Path, max_pages: int) -> str:
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    chunks: list[str] = []
    for page in reader.pages[:max_pages]:
        text = page.extract_text() or ""
        if text.strip():
            chunks.append(text)
    return "\n".join(chunks)


def extract_text_from_local_xlsx(path: Path, max_rows: int = 80) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    chunks: list[str] = []
    try:
        for sheet in workbook.worksheets[:3]:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_index >= max_rows:
                    break
                for value in row:
                    if value is None:
                        continue
                    text = str(value).strip()
                    if text:
                        chunks.append(text)
    finally:
        workbook.close()
    return "\n".join(chunks)


def extract_candidate_text(path: Path, extension: str) -> str:
    if extension == ".pdf":
        return extract_text_from_local_pdf(
            path,
            settings.PRESUPUESTO_CONTENT_CLASSIFICATION_MAX_PAGES,
        )
    if extension in {".xlsx", ".xls", ".xlsm"}:
        return extract_text_from_local_xlsx(path)
    return ""


def fetch_otro_presupuesto_candidates(portfolio_id: str) -> list[SecopDocumentDTO]:
    """SECOP documents not classified by filename that may be presupuesto by content."""
    candidates: list[SecopDocumentDTO] = []
    for document in fetch_all_documents_for_portfolio(portfolio_id):
        if document.document_type != DocumentType.OTRO:
            continue
        if is_archive_filename(document.file_name):
            continue
        extension = candidate_extension(document)
        if extension not in SUPPORTED_EXTENSIONS:
            continue
        if is_excluded_presupuesto_filename(document.file_name):
            continue
        candidates.append(document)

    def sort_key(item: SecopDocumentDTO) -> tuple[int, str]:
        extension = candidate_extension(item)
        priority = 0 if extension in {".xlsx", ".xls", ".xlsm"} else 1
        return priority, item.file_name.lower()

    return sorted(candidates, key=sort_key)


def tender_has_presupuesto(db: Session, tender: Tender) -> bool:
    return (
        db.query(TenderDocument.id)
        .filter(
            TenderDocument.tender_id == tender.id,
            TenderDocument.document_type == DocumentType.PRESUPUESTO.value,
        )
        .first()
        is not None
    )


@dataclass
class PresupuestoContentExtractionResult:
    candidates_inspected: int = 0
    documents_saved: int = 0
    documents_added: int = 0
    skipped_existing_presupuesto: bool = False
    errors: list[str] = field(default_factory=list)


def extract_presupuesto_by_content_for_tender(
    db: Session,
    tender: Tender,
    portfolio_id: str,
    storage: DocumentStorageService,
    upsert_document_record: Callable,
) -> PresupuestoContentExtractionResult:
    """Download and classify OTRO SECOP files by content when presupuesto is missing."""
    result = PresupuestoContentExtractionResult()
    if not settings.PRESUPUESTO_CONTENT_CLASSIFICATION_ENABLED:
        return result
    if tender_has_presupuesto(db, tender):
        result.skipped_existing_presupuesto = True
        return result

    candidates = fetch_otro_presupuesto_candidates(portfolio_id)
    if not candidates:
        return result

    staging_dir = Path(tempfile.mkdtemp(prefix="licitia_presupuesto_content_"))
    try:
        for candidate in candidates:
            if tender_has_presupuesto(db, tender):
                break

            extension = candidate_extension(candidate)
            staging_path = staging_dir / Path(candidate.file_name).name
            result.candidates_inspected += 1

            try:
                if not download_document_file(candidate, staging_path):
                    result.errors.append(f"download failed: {candidate.file_name}")
                    continue

                content_text = extract_candidate_text(staging_path, extension)
                if not classify_presupuesto_candidate(
                    candidate.file_name,
                    content_text,
                    staging_path,
                    extension,
                ):
                    continue

                presupuesto_document = candidate.model_copy(
                    update={"document_type": DocumentType.PRESUPUESTO}
                )
                object_key = build_document_object_key(
                    tender.external_id,
                    presupuesto_document.document_type,
                    presupuesto_document.file_name,
                    presupuesto_document.external_document_id,
                )
                destination = build_document_storage_path(
                    tender.external_id,
                    presupuesto_document.document_type,
                    presupuesto_document.file_name,
                    presupuesto_document.external_document_id,
                )
                destination.parent.mkdir(parents=True, exist_ok=True)
                shutil.copy2(staging_path, destination)
                storage.persist_local_file(destination, object_key)
                _, created = upsert_document_record(
                    db,
                    tender,
                    presupuesto_document,
                    object_key,
                )
                result.documents_saved += 1
                if created:
                    result.documents_added += 1
                logger.info(
                    "Presupuesto classified by content for tender %s: %s",
                    tender.external_id,
                    candidate.file_name,
                )
                break
            except Exception as exc:
                result.errors.append(f"{candidate.file_name}: {exc}")
                logger.warning(
                    "Presupuesto content classification failed for %s (%s): %s",
                    tender.external_id,
                    candidate.file_name,
                    exc,
                )
    finally:
        shutil.rmtree(staging_dir, ignore_errors=True)

    return result
