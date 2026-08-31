"""Extract plain text from stored tender documents (PDF, DOCX, Excel)."""
from __future__ import annotations

import re
import tempfile
import unicodedata
import zipfile
from contextlib import contextmanager
from pathlib import Path
from typing import Iterator

from app.core.logging import get_logger
from app.models.tender_document import TenderDocument
from app.services.document_storage import DocumentStorageService

logger = get_logger(__name__)


@contextmanager
def local_document_path(
    document: TenderDocument,
    storage: DocumentStorageService,
) -> Iterator[Path]:
    """Yield a local path to a stored document, downloading to temp if needed."""
    temp_dir: tempfile.TemporaryDirectory[str] | None = None
    try:
        local_path = storage.local_path(document.file_path)
        if local_path.is_file():
            yield local_path
            return

        temp_dir = tempfile.TemporaryDirectory(prefix="licitia_doc_")
        file_path = Path(temp_dir.name) / Path(document.file_name).name
        with file_path.open("wb") as handle:
            for chunk in storage.iter_file_chunks(document.file_path):
                handle.write(chunk)
        yield file_path
    finally:
        if temp_dir is not None:
            temp_dir.cleanup()


def _normalize_extracted_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKC", text or "")
    normalized = re.sub(r"[ \t]+\n", "\n", normalized)
    normalized = re.sub(r"\n{3,}", "\n\n", normalized)
    return normalized.strip()


def extract_docx_text(document: TenderDocument, storage: DocumentStorageService) -> str:
    """Read text from a DOCX file in storage."""
    try:
        with local_document_path(document, storage) as path:
            with zipfile.ZipFile(path) as archive:
                xml = archive.read("word/document.xml").decode("utf-8", errors="ignore")
    except Exception as exc:
        logger.warning("Failed to extract DOCX text from %s: %s", document.file_name, exc)
        return ""

    text = re.sub(r"<w:tab[^>]*/>", "\t", xml)
    text = re.sub(r"</w:p>", "\n", text)
    text = re.sub(r"<[^>]+>", "", text)
    return _normalize_extracted_text(text)


def extract_xlsx_text(document: TenderDocument, storage: DocumentStorageService) -> str:
    """Read cell values from the first sheets of an Excel workbook."""
    try:
        from openpyxl import load_workbook

        with local_document_path(document, storage) as path:
            workbook = load_workbook(path, read_only=True, data_only=True)
            chunks: list[str] = []
            for sheet_name in workbook.sheetnames[:3]:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(max_row=250, values_only=True):
                    values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                    if values:
                        chunks.append(" ".join(values))
            workbook.close()
            return _normalize_extracted_text("\n".join(chunks))
    except Exception as exc:
        logger.warning("Failed to extract XLSX text from %s: %s", document.file_name, exc)
        return ""


def extract_document_text(document: TenderDocument, storage: DocumentStorageService) -> str:
    """Best-effort text extraction for requirements and classification."""
    extension = (document.extension or Path(document.file_name).suffix).lower().lstrip(".")
    if extension == "pdf":
        from app.services.tender_summary.pdf_text import extract_pdf_text

        return extract_pdf_text(document, storage)
    if extension in {"docx", "doc"}:
        return extract_docx_text(document, storage)
    if extension in {"xlsx", "xls", "xlsm"}:
        return extract_xlsx_text(document, storage)
    return ""
