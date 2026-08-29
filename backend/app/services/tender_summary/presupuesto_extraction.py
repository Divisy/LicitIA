"""Extract AIU percentage from presupuesto XLSX (Formulario 1)."""
from __future__ import annotations

import re
import tempfile
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from app.core.logging import get_logger
from app.models.tender_document import TenderDocument
from app.services.document_storage import DocumentStorageService

logger = get_logger(__name__)


@dataclass
class PresupuestoExtraction:
    aiu_percentage: Optional[float] = None
    aiu_admin_percentage: Optional[float] = None
    aiu_imprevistos_percentage: Optional[float] = None
    aiu_utilidad_percentage: Optional[float] = None
    official_budget_total: Optional[float] = None
    budget_relation_note: Optional[str] = None


def extract_aiu_percentage_from_text(text: str) -> PresupuestoExtraction:
    """Extract AIU % from presupuesto text (PDF Formulario 1 or similar)."""
    result = PresupuestoExtraction()
    if not text:
        return result

    lowered = _normalize(text)

    direct = re.search(
        r"a\.?\s*i\.?\s*u\.?\s*[=:]\s*(\d{1,2}(?:[.,]\d+)?)\s*%",
        lowered,
    )
    if direct:
        result.aiu_percentage = float(direct.group(1).replace(",", "."))
        return normalize_aiu_percentages(_attach_aiu_components(result, lowered))

    admin = re.search(r"\ba\s*=\s*(\d{1,2}(?:[.,]\d+)?)\s*%", lowered)
    imprev = re.search(r"\bi\s*=\s*(\d{1,2}(?:[.,]\d+)?)\s*%", lowered)
    util = re.search(r"\bu\s*=\s*(\d{1,2}(?:[.,]\d+)?)\s*%", lowered)
    if admin and imprev and util:
        result.aiu_admin_percentage = float(admin.group(1).replace(",", "."))
        result.aiu_imprevistos_percentage = float(imprev.group(1).replace(",", "."))
        result.aiu_utilidad_percentage = float(util.group(1).replace(",", "."))
        result.aiu_percentage = round(
            result.aiu_admin_percentage
            + result.aiu_imprevistos_percentage
            + result.aiu_utilidad_percentage,
            2,
        )
        return normalize_aiu_percentages(result)

    percentages: list[float] = []
    for pattern in (
        r"total de administracion\s*\(a\)\s*(\d{1,2}(?:[.,]\d+)?)\s*%",
        r"imprevistos\s*\(i\)\s*(\d{1,2}(?:[.,]\d+)?)\s*%",
        r"utilidad\s*\(u\)\s*(\d{1,2}(?:[.,]\d+)?)\s*%",
    ):
        match = re.search(pattern, lowered)
        if match:
            percentages.append(float(match.group(1).replace(",", ".")))

    if len(percentages) >= 3:
        result.aiu_percentage = round(sum(percentages[:3]), 2)
        result.aiu_admin_percentage = percentages[0]
        result.aiu_imprevistos_percentage = percentages[1]
        result.aiu_utilidad_percentage = percentages[2]

    return normalize_aiu_percentages(result)


def _attach_aiu_components(result: PresupuestoExtraction, lowered: str) -> PresupuestoExtraction:
    admin = re.search(r"\ba\s*=\s*(\d{1,2}(?:[.,]\d+)?)\s*%", lowered)
    imprev = re.search(r"\bi\s*=\s*(\d{1,2}(?:[.,]\d+)?)\s*%", lowered)
    util = re.search(r"\bu\s*=\s*(\d{1,2}(?:[.,]\d+)?)\s*%", lowered)
    if admin:
        result.aiu_admin_percentage = float(admin.group(1).replace(",", "."))
    if imprev:
        result.aiu_imprevistos_percentage = float(imprev.group(1).replace(",", "."))
    if util:
        result.aiu_utilidad_percentage = float(util.group(1).replace(",", "."))
    return result


def format_aiu_display(extraction: PresupuestoExtraction) -> str:
    if extraction.aiu_percentage is None:
        return "No disponible"
    total = extraction.aiu_percentage
    if (
        extraction.aiu_admin_percentage is not None
        and extraction.aiu_imprevistos_percentage is not None
        and extraction.aiu_utilidad_percentage is not None
    ):
        return (
            f"{total:.2f}% "
            f"(A {extraction.aiu_admin_percentage:g}% · "
            f"I {extraction.aiu_imprevistos_percentage:g}% · "
            f"U {extraction.aiu_utilidad_percentage:g}%)"
        )
    return f"{total:.2f}%"


_AIU_CONTEXT_RE = re.compile(
    r"a\.?\s*i\.?\s*u\.?\s*[=:(]|\badministraci[oó]n\s*\(a\b|\bimprevistos\s*\(i\b|\butilidad\s*\(u\b",
    re.IGNORECASE,
)
_REJECT_AIU_EVIDENCE_RE = re.compile(
    r"\bexperiencia\b|\banticipo\b|\binterventor",
    re.IGNORECASE,
)


def has_presupuesto_aiu_context(text: str) -> bool:
    """True when text likely comes from a presupuesto AIU block, not pliego experiencia."""
    if not text or not text.strip():
        return False
    lowered = _normalize(text)
    if re.search(r"\bexperiencia\b", lowered) and not _AIU_CONTEXT_RE.search(lowered):
        return False
    if _AIU_CONTEXT_RE.search(lowered):
        return True
    if re.search(r"\ba\s*=\s*\d", lowered) and re.search(r"\bi\s*=\s*\d", lowered):
        return True
    return False


def is_plausible_aiu_range(parsed: PresupuestoExtraction) -> bool:
    """Typical obra pública AIU in Colombia: total ~15–40%, A ~18–30%, I ~1–3%, U ~3–8%."""
    total = parsed.aiu_percentage
    if total is None or total < 10 or total > 50:
        return False

    admin = parsed.aiu_admin_percentage
    imprev = parsed.aiu_imprevistos_percentage
    util = parsed.aiu_utilidad_percentage

    if admin is not None and (admin < 10 or admin > 35):
        return False
    if imprev is not None and (imprev < 0.5 or imprev > 10):
        return False
    if util is not None and (util < 2 or util > 15):
        return False
    return True


def _scale_decimal_percent(value: float) -> float:
    if 0 < value < 5:
        return round(value * 100, 2)
    return value


def normalize_aiu_percentages(parsed: PresupuestoExtraction) -> PresupuestoExtraction:
    """Fix common OCR/vision misreads like 0.24 instead of 24%."""
    if parsed.aiu_percentage is None:
        return parsed

    admin = parsed.aiu_admin_percentage
    imprev = parsed.aiu_imprevistos_percentage
    util = parsed.aiu_utilidad_percentage

    if admin is not None and imprev is not None and util is not None:
        if max(admin, imprev, util) < 5 and (admin + imprev + util) < 5:
            admin = _scale_decimal_percent(admin)
            imprev = _scale_decimal_percent(imprev)
            util = _scale_decimal_percent(util)
            return PresupuestoExtraction(
                aiu_percentage=round(admin + imprev + util, 2),
                aiu_admin_percentage=admin,
                aiu_imprevistos_percentage=imprev,
                aiu_utilidad_percentage=util,
                official_budget_total=parsed.official_budget_total,
                budget_relation_note=parsed.budget_relation_note,
            )

    total = parsed.aiu_percentage
    if 0 < total < 5:
        scaled_total = round(total * 100, 2)
        if 15 <= scaled_total <= 50:
            return PresupuestoExtraction(
                aiu_percentage=scaled_total,
                aiu_admin_percentage=admin,
                aiu_imprevistos_percentage=imprev,
                aiu_utilidad_percentage=util,
                official_budget_total=parsed.official_budget_total,
                budget_relation_note=parsed.budget_relation_note,
            )

    return parsed


def is_credible_aiu_extraction(
    parsed: PresupuestoExtraction,
    *,
    evidence: Optional[str] = None,
) -> bool:
    if parsed.aiu_percentage is None:
        return False

    parsed = normalize_aiu_percentages(parsed)
    if not is_plausible_aiu_range(parsed):
        return False

    components = (
        parsed.aiu_admin_percentage,
        parsed.aiu_imprevistos_percentage,
        parsed.aiu_utilidad_percentage,
    )
    if all(value is not None for value in components):
        expected = round(sum(components), 2)  # type: ignore[type-arg]
        if abs(expected - parsed.aiu_percentage) > 0.15:
            return False
        return True

    evidence_text = (evidence or "").strip()
    if evidence_text:
        if _REJECT_AIU_EVIDENCE_RE.search(evidence_text) and not _AIU_CONTEXT_RE.search(
            _normalize(evidence_text)
        ):
            return False
        if _AIU_CONTEXT_RE.search(_normalize(evidence_text)):
            return True

    return False


def _normalize(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().lower()


def resolve_official_budget_total(
    secop_amount: Optional[float],
    extracted_amount: Optional[float],
    *,
    min_ratio: float = 0.25,
    max_ratio: float = 4.0,
) -> Optional[float]:
    """Prefer SECOP when presupuesto extraction returns an implausible total."""
    if extracted_amount is None or extracted_amount <= 0:
        if secop_amount is None or secop_amount <= 0:
            return None
        return float(secop_amount)
    if secop_amount is None or secop_amount <= 0:
        return float(extracted_amount)

    secop = float(secop_amount)
    extracted = float(extracted_amount)
    ratio = extracted / secop
    if ratio < min_ratio or ratio > max_ratio:
        return secop
    return extracted


def _parse_number(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    cleaned = text.replace("$", "").replace(" ", "").replace("\u00a0", "")
    if re.fullmatch(r"\d{1,3}(\.\d{3})+", cleaned):
        cleaned = cleaned.replace(".", "")
    elif re.fullmatch(r"\d{1,3}(,\d{3})+", cleaned):
        cleaned = cleaned.replace(",", "")
    elif "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    elif "." in cleaned:
        whole, fraction = cleaned.split(".", 1)
        if whole.isdigit() and fraction.isdigit() and len(fraction) == 3:
            cleaned = whole + fraction
    cleaned = re.sub(r"[^\d.-]", "", cleaned)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def extract_from_presupuesto_xlsx(
    document: TenderDocument,
    storage: DocumentStorageService,
) -> PresupuestoExtraction:
    result = PresupuestoExtraction()
    extension = (document.extension or "").lower()
    if extension not in {"xlsx", "xls", "xlsm"}:
        return result

    temp_dir: tempfile.TemporaryDirectory[str] | None = None
    try:
        local_path = storage.local_path(document.file_path)
        if local_path.is_file():
            workbook_path = local_path
        else:
            temp_dir = tempfile.TemporaryDirectory(prefix="licitia_xlsx_")
            workbook_path = Path(temp_dir.name) / Path(document.file_name).name
            with workbook_path.open("wb") as handle:
                for chunk in storage.iter_file_chunks(document.file_path):
                    handle.write(chunk)

        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        percentages: list[float] = []
        totals: list[float] = []

        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [cell for cell in row if cell is not None]
                if not cells:
                    continue
                labels = [_normalize(str(cell)) for cell in cells[:4] if isinstance(cell, str)]
                joined = " ".join(labels)

                if any(token in joined for token in ("administracion", "impuestos", "utilidad", "aiu")):
                    for cell in cells[1:6]:
                        number = _parse_number(cell)
                        if number is not None and 0 < number <= 100:
                            percentages.append(number)

                if "presupuesto oficial" in joined or "valor total" in joined or joined.startswith("total"):
                    for cell in reversed(cells):
                        number = _parse_number(cell)
                        if number is not None and number > 1000:
                            totals.append(number)
                            break

        workbook.close()

        if percentages:
            if len(percentages) >= 3:
                result.aiu_percentage = round(sum(percentages[:3]), 2)
                result.aiu_admin_percentage = round(percentages[0], 2)
                result.aiu_imprevistos_percentage = round(percentages[1], 2)
                result.aiu_utilidad_percentage = round(percentages[2], 2)
            else:
                result.aiu_percentage = round(max(percentages), 2)

        if totals:
            result.official_budget_total = max(totals)

        if result.official_budget_total:
            result.budget_relation_note = "Presupuesto oficial identificado en Formulario 1"

        return result
    except Exception as exc:
        logger.warning("Failed to parse presupuesto XLSX %s: %s", document.file_name, exc)
        return result
    finally:
        if temp_dir is not None:
            temp_dir.cleanup()
